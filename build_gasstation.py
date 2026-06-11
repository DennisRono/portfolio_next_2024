from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── Color palette ──────────────────────────────────────────────────
HDR_BG   = "1F3864"   # dark navy
HDR_FG   = "FFFFFF"
SUB_BG   = "2E75B6"   # medium blue
SUB_FG   = "FFFFFF"
CAT_BG   = "BDD7EE"   # light blue
CAT_FG   = "1F3864"
ALT_BG   = "EBF3FB"   # very light blue
WHITE    = "FFFFFF"
GREEN    = "E2EFDA"
GOLD     = "FFF2CC"
ORANGE   = "FCE4D6"

def hdr_font(sz=11, bold=True, color=HDR_FG):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def cell_font(sz=10, bold=False, color="000000"):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="AAAAAA")
thick= Side(style="medium", color="1F3864")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def style_header_row(ws, row, cols, bg=HDR_BG, fg=HDR_FG, sz=11):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = hdr_font(sz=sz, color=fg)
        cell.alignment = center()
        cell.border = thin_border()

def style_subheader(ws, row, cols, bg=SUB_BG, fg=SUB_FG):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = hdr_font(sz=10, color=fg)
        cell.alignment = center()
        cell.border = thin_border()

def style_category(ws, row, cols, bg=CAT_BG):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = hdr_font(sz=10, color=CAT_FG)
        cell.alignment = left()
        cell.border = thin_border()

def style_data_row(ws, row, cols, alt=False):
    bg = ALT_BG if alt else WHITE
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = cell_font()
        if c == 1:
            cell.alignment = left()
        else:
            cell.alignment = center()
        cell.border = thin_border()

def write_title(ws, title, subtitle, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    t = ws.cell(row=1, column=1, value=title)
    t.fill = fill(HDR_BG); t.font = hdr_font(sz=14); t.alignment = center()
    s = ws.cell(row=2, column=1, value=subtitle)
    s.fill = fill(SUB_BG); s.font = hdr_font(sz=10); s.alignment = center()
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_section(ws, cur_row, category_label, cols, items):
    """items = list of tuples: (name, *prices, notes)"""
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=cols)
    ws.cell(row=cur_row, column=1, value=f"  ▶  {category_label}")
    style_category(ws, cur_row, cols)
    ws.row_dimensions[cur_row].height = 16
    cur_row += 1
    for i, item in enumerate(items):
        for j, val in enumerate(item, 1):
            ws.cell(row=cur_row, column=j, value=val)
        style_data_row(ws, cur_row, cols, alt=(i % 2 == 1))
        ws.row_dimensions[cur_row].height = 15
        cur_row += 1
    return cur_row + 1  # blank gap row

# ════════════════════════════════════════════════════════════════════
#  SHEET 1 – BEER
# ════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "🍺 Beer"
COLS = 8
write_title(ws, "BEER PRICE LIST", "All sizes | Prices subject to change", COLS)
set_col_widths(ws, [26, 9, 9, 9, 11, 11, 11, 18])

headers = ["Brand / Product", "Single\n16oz", "4-Pack", "6-Pack",
           "12-Pack", "24-Pack", "24oz Single", "Notes"]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header_row(ws, 3, COLS)
ws.row_dimensions[3].height = 32

cur = 4
beer_data = [
    ("Bud Ice",          "$1.29", "",      "",       "$11.99", "$29.99", "$2.39", "8pk $9.99"),
    ("Bud Light",        "$1.39", "$5.29", "$9.99",  "$14.99", "",       "$2.49", ""),
    ("Bud Light Lime",   "",      "",      "$8.99",  "",       "",       "",      ""),
    ("Bud Light Orange", "",      "",      "",       "",       "",       "$2.29", ""),
    ("Bud Light Platinum","",     "",      "$9.99",  "$19.49", "",       "",      ""),
    ("Bud Light Clamato","",      "",      "",       "",       "",       "$2.59", ""),
    ("Budweiser",        "$1.49", "$5.49", "$9.49",  "$14.99", "",       "$2.49", ""),
    ("Colt 45",          "",      "",      "",       "",       "",       "$1.99", ""),
    ("Coors Banquet",    "$1.39", "$4.49", "",       "",       "",       "",      ""),
    ("Coors Light",      "$1.49", "$5.49", "$10.99", "$14.99", "",       "$1.99", "8oz 12pk $6.99"),
    ("Corona",           "$1.49", "$8.99", "$9.99",  "$19.49", "",       "$3.29", "7oz 6pk $10.49"),
    ("Coronita 24pk",    "",      "",      "",       "",       "$29.99", "",      ""),
    ("Dos Equis XXX",    "$1.49", "$5.99", "$9.99",  "",       "",       "$2.99", ""),
    ("Four LOKO",        "",      "",      "",       "",       "",       "$3.99", ""),
    ("Guinness",         "$2.49", "$9.49", "$9.49",  "",       "",       "",      "Draught $9.99 / Extra Stout $9.49"),
    ("Heineken",         "$2.49", "$9.99", "$8.99",  "$18.99", "",       "$2.99", ""),
    ("High Life",        "",      "",      "$8.99",  "$11.49", "",       "$2.99", ""),
    ("Hurricane",        "$1.99", "$6.99", "",       "",       "",       "$1.79", ""),
    ("Ice House",        "$1.29", "$4.99", "",       "",       "",       "",      ""),
    ("Jack Daniel's",    "",      "",      "$9.99",  "",       "",       "",      ""),
    ("Kinky",            "",      "",      "$6.99",  "",       "",       "",      ""),
    ("Lime-A-Rita",      "",      "",      "",       "",       "",       "$2.99", ""),
    ("Michelob Ultra 4pk","$1.29","$4.99", "",       "",       "",       "",      ""),
    ("Michelob",         "",      "",      "$10.99", "$18.99", "",       "$2.99", "Alum 12pk $20.99"),
    ("Michelob 8oz",     "",      "",      "",       "$13.49", "",       "",      ""),
    ("Mickey's",         "",      "",      "$9.99",  "",       "",       "",      ""),
    ("Mike's Hard",      "",      "",      "$6.99",  "",       "",       "",      ""),
    ("Miller High Life", "$1.99", "$6.99", "",       "$11.49", "",       "$2.99", ""),
    ("Miller Light",     "$1.29", "$5.49", "",       "$11.49", "",       "",      ""),
    ("Modelo",           "$1.99", "$6.99", "$10.99", "$19.49", "",       "$3.29", "Modelo Cans 12pk $19.99"),
    ("Natural Ice",      "$2.49", "$9.49", "",       "",       "",       "$2.59", ""),
    ("Natural Light",    "$1.39", "$4.99", "",       "$12.99", "",       "$1.99", "15pk $12.99"),
    ("Old English 800",  "$1.39", "$4.99", "",       "",       "",       "",      ""),
    ("Pabst Blue Ribbon","$1.39", "$4.99", "",       "",       "",       "",      ""),
    ("Pacifico",         "",      "",      "$11.49", "",       "",       "",      ""),
    ("Red Stripe",       "$1.79", "$6.49", "$9.49",  "",       "",       "",      ""),
    ("Reds",             "",      "",      "$10.49", "",       "",       "",      ""),
    ("Seagram's Spike",  "",      "$5.49", "",       "",       "",       "$2.49", ""),
    ("Smirnoff Ice",     "",      "",      "$10.49", "",       "",       "$2.39", ""),
    ("Steel Reserve",    "$1.29", "$4.99", "",       "",       "",       "$1.99", ""),
    ("Stella Artois",    "",      "",      "$10.99", "",       "",       "",      ""),
    ("Victoria",         "",      "",      "$10.99", "",       "",       "",      ""),
    ("White Claw",       "",      "",      "",       "",       "",       "$2.99", ""),
    ("Angry Orchard",    "",      "",      "$9.99",  "",       "",       "",      "Single $5.99"),
    ("Clubtails",        "",      "",      "$8.49",  "",       "",       "",      ""),
    ("Buzz Ball",        "",      "",      "",       "",       "",       "$3.49", "Loose"),
    ("Club Tails",       "",      "",      "",       "",       "",       "",      "Loose"),
    ("Boot Legger",      "$1.99", "",      "",       "",       "",       "",      ""),
]

bottle_items = [
    ("Bud Light 18oz Bottle", "", "", "", "", "", "$1.99", ""),
    ("Bud Light Platinum Bottle", "", "", "", "", "", "$2.49", ""),
    ("Corona Bottle", "", "", "", "", "", "$3.29", ""),
    ("Heineken Bottle", "", "", "", "", "", "$2.99", ""),
    ("Modelo Bottle", "", "", "", "", "", "$3.29", ""),
    ("Smirnoff Bottle 12oz", "", "", "", "", "", "$3.19", ""),
]

cur = add_section(ws, cur, "CANNED BEER", COLS, beer_data)
cur = add_section(ws, cur, "BOTTLED BEER", COLS, bottle_items)
ws.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 2 – WINE & MALT BEVERAGES
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🍷 Wine & Malt")
COLS2 = 5
write_title(ws2, "WINE & MALT BEVERAGES", "All sizes | Prices subject to change", COLS2)
set_col_widths(ws2, [28, 10, 10, 10, 20])
hdrs2 = ["Brand / Product", "Price", "Size", "Alt Price", "Notes"]
for j, h in enumerate(hdrs2, 1):
    ws2.cell(row=3, column=j, value=h)
style_header_row(ws2, 3, COLS2)
ws2.row_dimensions[3].height = 22

wine_data = [
    ("Andre Champagne",         "$6.99",  "750ml",  "",       ""),
    ("Arbor Mist",               "$5.49",  "750ml",  "",       ""),
    ("Barefoot",                 "$1.79",  "187ml",  "$6.99",  "4pk $6.99 | 750ml $8.99"),
    ("Barefoot Bubbly",          "$11.99", "750ml",  "",       ""),
    ("Boone's Farm",             "$4.99",  "",       "",       ""),
    ("Boot Legger Wine",         "$1.99",  "",       "",       ""),
    ("Buzz Ball",                "$3.49",  "",       "",       ""),
    ("Carlo Rossi",              "$3.99",  "",       "",       ""),
    ("Daily's Pouches",          "$1.99",  "",       "",       ""),
    ("Pre Game",                 "$1.99",  "",       "",       ""),
    ("Gallo",                    "$3.99",  "750ml",  "$7.99",  "1.75L $7.99"),
    ("Liberty Creek",            "$7.99",  "1.75L",  "",       ""),
    ("MD 20/20 (MD 2020)",       "$1.99",  "375ml",  "$3.99",  "750ml $3.99"),
    ("Moscato",                  "$7.99",  "",       "",       ""),
    ("MYX",                      "$12.99", "",       "",       ""),
    ("La Rosa",                  "$10.99", "750ml",  "",       ""),
    ("Rosé 1.5L",                "$25.99", "1.5L",   "",       ""),
    ("Peter Vella (Peter Home)", "$1.79",  "187ml",  "$6.99",  "4pk $6.99 | 750ml $5.99 | 1.75L $9.99"),
    ("Taylor Black",             "$10.99", "",       "",       ""),
    ("Taylor Port",              "$5.99",  "375ml",  "$9.99",  "750ml; Big $17.99"),
    ("Verdi",                    "$7.99",  "",       "",       ""),
    ("Richards Wild Irish Rose", "$3.99",  "375ml",  "$4.00",  "750ml"),
    ("Cocktail (Cocktail Tail)", "$6.99",  "750ml",  "",       ""),
]

malt_data = [
    ("Seagram's Escapes 4pk",   "$5.49",  "4pk",    "",       ""),
    ("Four LOKO",                "$3.99",  "24oz",   "",       ""),
    ("Kinky",                    "$6.99",  "6pk",    "",       ""),
    ("Smirnoff Ice",             "$2.39",  "24oz",   "",       ""),
    ("White Claw",               "$2.99",  "24oz",   "",       ""),
    ("Lime-A-Rita",              "$2.99",  "24oz",   "",       ""),
    ("Mike's Hard",              "$6.99",  "6pk",    "",       ""),
    ("Hurricane",                "$1.79",  "24oz",   "",       ""),
]

cur2 = 4
cur2 = add_section(ws2, cur2, "WINE", COLS2, wine_data)
cur2 = add_section(ws2, cur2, "MALT BEVERAGES / ALCOPOPS", COLS2, malt_data)
ws2.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 3 – LIQUOR (prices missing from original images)
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🥃 Liquor")
COLS3 = 7
write_title(ws3, "LIQUOR PRICE LIST", "Sizes: Nip(50ml) | Half Pint(200ml) | Pint(375ml) | Fifth(750ml) | Half Gallon(1.75L)", COLS3)
set_col_widths(ws3, [30, 9, 10, 10, 10, 10, 18])
ws3.row_dimensions[2].height = 22
hdrs3 = ["Brand / Product", "Nip\n50ml", "Half Pint\n200ml", "Pint\n375ml", "Fifth\n750ml", "Half Gal\n1.75L", "Notes"]
for j, h in enumerate(hdrs3, 1):
    ws3.cell(row=3, column=j, value=h)
style_header_row(ws3, 3, COLS3)
ws3.row_dimensions[3].height = 36

vodka_data = [
    ("Absolut Vodka", "", "", "", "", "", ""), ("Absolut Citron", "", "", "", "", "", ""),
    ("Absolut Apeach", "", "", "", "", "", ""), ("Absolut Lime", "", "", "", "", "", ""),
    ("Aristocrat Vodka", "", "", "", "", "", "Budget brand"), ("Barton Vodka", "", "", "", "", "", "Premium label"),
    ("Belvedere", "", "", "", "", "", "Premium"), ("Burnett's 86 Proof", "", "", "", "", "", ""),
    ("Ciroc (Original)", "", "", "", "", "", ""), ("Ciroc Apple", "", "", "", "", "", ""),
    ("Ciroc Berry", "", "", "", "", "", ""), ("Ciroc Coconut", "", "", "", "", "", ""),
    ("Ciroc Mango", "", "", "", "", "", ""), ("Ciroc Peach", "", "", "", "", "", ""),
    ("Ciroc Pineapple", "", "", "", "", "", ""), ("Ciroc Red Berry", "", "", "", "", "", ""),
    ("Ciroc Summer Citrus", "", "", "", "", "", ""), ("Ciroc Watermelon", "", "", "", "", "", ""),
    ("Fris Vodka", "", "", "", "", "", ""), ("Grey Goose", "", "", "", "", "", "Premium"),
    ("Ketel One", "", "", "", "", "", ""), ("New Amsterdam (Original)", "", "", "", "", "", ""),
    ("New Amsterdam Apple", "", "", "", "", "", ""), ("New Amsterdam Mango", "", "", "", "", "", ""),
    ("New Amsterdam Peach", "", "", "", "", "", ""), ("New Amsterdam Pineapple", "", "", "", "", "", ""),
    ("New Amsterdam Raspberry", "", "", "", "", "", ""), ("New Amsterdam Watermelon", "", "", "", "", "", ""),
    ("Pinnacle Vodka", "", "", "", "", "", ""), ("Pinnacle Citrus", "", "", "", "", "", ""),
    ("Pinnacle Grape", "", "", "", "", "", ""), ("Pinnacle Mango", "", "", "", "", "", ""),
    ("Pinnacle Peach", "", "", "", "", "", ""), ("Pinnacle Raspberry", "", "", "", "", "", ""),
    ("Pinnacle Whipped", "", "", "", "", "", ""), ("Platinum 10X Vodka", "", "", "", "", "", ""),
    ("Seagram's Extra Smooth", "", "", "", "", "", ""), ("Skyy Vodka", "", "", "", "", "", ""),
    ("Skyy Mango Pineapple", "", "", "", "", "", ""), ("Skol Premium", "", "", "", "", "", "Budget brand"),
    ("Smirnoff No. 21", "", "", "", "", "", ""), ("Smirnoff 100 Proof", "", "", "", "", "", ""),
    ("Smirnoff Blue Raspberry", "", "", "", "", "", ""), ("Smirnoff Green Apple", "", "", "", "", "", ""),
    ("Smirnoff Lemonade", "", "", "", "", "", ""), ("Smirnoff Peach", "", "", "", "", "", ""),
    ("Smirnoff Pineapple", "", "", "", "", "", ""), ("Smirnoff Strawberry", "", "", "", "", "", ""),
    ("Svedka Vodka", "", "", "", "", "", ""), ("Svedka Mango Pineapple", "", "", "", "", "", ""),
    ("Svedka Strawberry Lemonade", "", "", "", "", "", ""), ("Taaka Vodka 80", "", "", "", "", "", "Budget brand"),
    ("Tito's Handmade Vodka", "", "", "", "", "", ""),
]

whiskey_data = [
    ("Crown Royal Fine De Luxe", "", "", "", "", "", ""), ("Crown Royal Apple", "", "", "", "", "", ""),
    ("Crown Royal Vanilla", "", "", "", "", "", ""), ("Crown Royal Peach", "", "", "", "", "", ""),
    ("Crown Royal Blackberry", "", "", "", "", "", ""), ("Crown Royal Black", "", "", "", "", "", ""),
    ("Crown Royal Reserve", "", "", "", "", "", ""), ("Evan Williams", "", "", "", "", "", "Kentucky Bourbon"),
    ("Gentleman Jack", "", "", "", "", "", "Tennessee Whiskey"), ("Jack Daniel's No. 7", "", "", "", "", "", "Tennessee Whiskey"),
    ("Jack Daniel's Tennessee Honey", "", "", "", "", "", ""), ("Jack Daniel's Tennessee Fire", "", "", "", "", "", ""),
    ("Jack Daniel's Single Barrel", "", "", "", "", "", ""), ("Jim Beam White", "", "", "", "", "", ""),
    ("Jim Beam Fire", "", "", "", "", "", ""), ("Jim Beam Apple", "", "", "", "", "", ""),
    ("Jim Beam Peach", "", "", "", "", "", ""), ("Jim Beam Vanilla", "", "", "", "", "", ""),
    ("Jim Beam Honey", "", "", "", "", "", ""), ("Jim Beam Devil's Cut", "", "", "", "", "", ""),
    ("Jim Beam Black", "", "", "", "", "", ""), ("Knob Creek", "", "", "", "", "", "Small Batch Bourbon"),
    ("Maker's Mark", "", "", "", "", "", ""), ("Maker's Mark Honey", "", "", "", "", "", ""),
    ("Seagram's 7 Crown", "", "", "", "", "", "American Blended"), ("Seagram's VO Canadian", "", "", "", "", "", ""),
    ("Woodford Reserve", "", "", "", "", "", "Premium Bourbon"), ("Canadian Mist", "", "", "", "", "", ""),
    ("Club 400 Whiskey", "", "", "", "", "", "Value brand"), ("Old Forester", "", "", "", "", "", ""),
    ("R&R Rich & Rare", "", "", "", "", "", ""),
]

cognac_data = [
    ("Courvoisier VS", "", "", "", "", "", ""), ("Courvoisier VSOP", "", "", "", "", "", ""),
    ("D'Ussé VSOP", "", "", "", "", "", ""), ("E&J VS Brandy", "", "", "", "", "", ""),
    ("Hennessy VS", "", "", "", "", "", ""), ("Hennessy VSOP", "", "", "", "", "", ""),
    ("Paul Masson Amber", "", "", "", "", "", "Grande Amber Brandy"), ("Paul Masson Peach", "", "", "", "", "", ""),
    ("Rémy Martin VSOP", "", "", "", "", "", ""),
]

rum_data = [("Bacardi Superior", "", "", "", "", "", ""), ("Captain Morgan", "", "", "", "", "", ""), ("Malibu Coconut", "", "", "", "", "", "")]
tequila_data = [("Jose Cuervo Gold", "", "", "", "", "", ""), ("Jose Cuervo Silver", "", "", "", "", "", ""),
                ("Lunazul Blanco", "", "", "", "", "", ""), ("Lunazul Reposado", "", "", "", "", "", ""),
                ("Patron Silver", "", "", "", "", "", "Premium")]
gin_data = [("Barton Gin", "", "", "", "", "", "Value brand"), ("Burnett's Gin", "", "", "", "", "", ""),
            ("Seagram's Gin", "", "", "", "", "", ""), ("Tanqueray", "", "", "", "", "", "")]

cur3 = 4
cur3 = add_section(ws3, cur3, "VODKA", COLS3, vodka_data)
cur3 = add_section(ws3, cur3, "WHISKEY / BOURBON / CANADIAN", COLS3, whiskey_data)
cur3 = add_section(ws3, cur3, "COGNAC & BRANDY", COLS3, cognac_data)
cur3 = add_section(ws3, cur3, "RUM", COLS3, rum_data)
cur3 = add_section(ws3, cur3, "TEQUILA", COLS3, tequila_data)
cur3 = add_section(ws3, cur3, "GIN", COLS3, gin_data)
ws3.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 4 – TOBACCO
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🚬 Tobacco")
COLS4 = 5
write_title(ws4, "TOBACCO & CIGARS", "Cigarettes | Cigars | Wraps | Pouches", COLS4)
set_col_widths(ws4, [32, 10, 10, 10, 22])
hdrs4 = ["Product", "Single", "Box/Pack", "Carton", "Notes"]
for j, h in enumerate(hdrs4, 1):
    ws4.cell(row=3, column=j, value=h)
style_header_row(ws4, 3, COLS4)
ws4.row_dimensions[3].height = 22

cigs_data = [
    ("Newport 100s Box", "", "", "", "Full flavor menthol"), ("Newport 100s Soft Pack", "", "", "", ""),
    ("Newport Short (Regular) Soft", "", "", "", ""), ("Newport Smooth Select", "", "", "", ""),
    ("Newport Menthol Gold", "", "", "", ""), ("Marlboro Red Box", "", "", "", ""),
    ("Marlboro Light Box", "", "", "", ""), ("Marlboro Menthol", "", "", "", ""),
    ("Marlboro Gold", "", "", "", ""), ("Seneca (Regular)", "", "", "", "Value brand"),
    ("Seneca Menthol", "", "", "", ""), ("Pall Mall", "", "", "", ""), ("Camel", "", "", "", ""),
    ("Kool", "", "", "", ""),
]

cigars_data = [
    ("Black & Mild Original", "$0.99", "", "", ""), ("Black & Mild Wine", "$0.99", "", "", ""),
    ("Black & Mild Cream", "$0.99", "", "", ""), ("Black & Mild Jazz", "$0.99", "", "", ""),
    ("Black & Mild Casino", "$0.99", "", "", ""), ("Black & Mild Wood Tip", "$1.29", "", "", ""),
    ("Black & Mild Apple", "$0.99", "", "", ""), ("Black & Mild Mild (Box)", "", "$5.99", "", ""),
    ("Swisher Sweets Original", "$0.99", "$1.79", "", "2-for-$1"), ("Swisher Sweets Grape", "$0.99", "$1.79", "", ""),
    ("Swisher Sweets Strawberry", "$0.99", "$1.79", "", ""), ("Swisher Sweets Peach", "$0.99", "$1.79", "", ""),
    ("Swisher Sweets Blueberry", "$0.99", "", "", ""), ("Game (Original)", "$0.99", "$1.79", "", ""),
    ("Game Honey", "$0.99", "", "", ""), ("Game Peach", "$0.99", "", "", ""),
    ("Leaf (Natural)", "$0.99", "", "", ""), ("Leaf Grape", "$0.99", "", "", ""),
    ("4K Wraps (Backwoods style)", "$1.49", "", "", ""), ("4K Cigars", "$0.99", "", "", ""),
    ("White Owl", "$0.99", "", "", ""), ("Dutch Masters", "$0.99", "", "", ""),
    ("Optimo", "$0.99", "", "", ""),
]

dip_data = [
    ("Grizzly Wintergreen", "", "", "", ""), ("Grizzly Mint", "", "", "", ""),
    ("Copenhagen", "", "", "", ""), ("Skoal", "", "", "", ""),
    ("Zyn Nicotine Pouches", "", "", "", "Various strengths"),
]

cur4 = 4
cur4 = add_section(ws4, cur4, "CIGARETTES", COLS4, cigs_data)
cur4 = add_section(ws4, cur4, "CIGARS / BLUNTS / WRAPS", COLS4, cigars_data)
cur4 = add_section(ws4, cur4, "DIP / POUCHES", COLS4, dip_data)
ws4.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 5 – SNACKS
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🍿 Snacks")
COLS5 = 5
write_title(ws5, "SNACKS & CHIPS", "Chips | Crackers | Cookies | Candy | Pork Skins", COLS5)
set_col_widths(ws5, [32, 10, 10, 12, 22])
hdrs5 = ["Product", "Price", "Deal Price", "Deal Qty", "Notes"]
for j, h in enumerate(hdrs5, 1):
    ws5.cell(row=3, column=j, value=h)
style_header_row(ws5, 3, COLS5)
ws5.row_dimensions[3].height = 22

gf_data = [
    ("Golden Flake Hot Chips", "$0.65", "", "", "Small bag"), ("Golden Flake Mesquite BBQ", "$0.65", "", "", "Dip Style"),
    ("Golden Flake Sour Cream & Onion", "$0.65", "", "", ""), ("Golden Flake Vinegar & Salt", "$0.65", "", "", ""),
    ("Golden Flake Cheese Puffs", "$0.65", "", "", ""), ("Golden Flake Jalapeño Cheddar", "$0.65", "", "", ""),
    ("Golden Flake Louisiana Hot Sauce PS", "$0.65", "", "", "Pork Skins"), ("Golden Flake Old Fashioned Pork Skins", "$0.65", "", "", ""),
    ("Golden Flake Super Stripe Red Pepper", "$0.65", "", "", ""), ("Golden Flake Trail Mix", "$2.99", "", "", "Big bag"),
    ("Golden Flake Peanuts", "$0.65", "", "", ""), ("Golden Flake Big Bag", "$2.99", "", "", ""),
]

chips_data = [
    ("Herr's Salt & Pepper", "$2.99", "", "", ""), ("Herr's Sour Cream French Onion", "$2.99", "", "", "New Flavor"),
    ("Herr's Creamy Dill", "$2.99", "", "", ""), ("Herr's Jalapeño Poppers", "$2.99", "", "", "Cheesy Popcorn"),
    ("Herr's Heat", "$2.99", "", "", ""), ("O-Ke-Doke Cheese Popcorn", "$2.99", "", "", ""),
    ("O-Ke-Doke Hot Stuff", "$2.99", "", "", ""), ("Cheez-It Original", "$1.69", "", "", "Small bag"),
    ("Cheez-It Grooves Cheddar Ranch", "$2.99", "", "", ""), ("Cheez-It Puff'd Double Cheese", "$2.99", "", "", ""),
    ("Cheez-It Snap'd", "$2.99", "", "", ""), ("Cheez-It Extra Cheesy", "$2.99", "", "", ""),
    ("Cheez-It Spicy", "$2.99", "", "", ""), ("Cheez-It Duoz", "$2.99", "", "", ""),
    ("Ritz Toasted Chips Sour Cream", "$2.99", "", "", ""), ("Andy Capp's Ranch Fries", "$1.69", "", "", ""),
    ("Andy Capp's Onion Rings", "$1.69", "", "", ""), ("Andy Capp's Hot Fries", "$1.69", "", "", ""),
    ("Andy Capp's Hot Onion Rings", "$1.69", "", "", ""), ("Snyder's Mini Pretzels", "$1.69", "", "", ""),
    ("Snyder's Buffalo Wing", "$1.69", "", "", ""), ("Snyder's Pretzel Dips White Choc", "$2.99", "", "", ""),
    ("Snyder's Pieces Honey Mustard", "$2.99", "", "", ""), ("Pringles Mingles Cheddar/Sour Cream", "$2.99", "", "", "New! Light & Crispy"),
    ("Vlasic Pickle Balls Dill Pickle", "$2.99", "", "", "Corn Puffs"), ("Chex Mix Original", "$2.99", "", "", ""),
    ("Chex Mix Spicy", "$2.99", "", "", ""), ("Chex Mix Buffalo", "$2.99", "", "", ""),
    ("Apple Jacks Jumbo Snax", "$2.99", "", "", ""), ("Gardetto's Special Request", "$2.99", "", "", "Garlic Rye Chips"),
    ("Doritos", "$1.69", "", "", ""), ("Fritos", "$0.99", "", "", ""), ("Lay's Classic", "$1.69", "", "", ""),
]

cookies_data = [
    ("Oreo Original", "$1.69", "$2.99", "Big Bag", ""), ("Oreo Golden Double Stuf", "$2.99", "", "", "King Size"),
    ("Oreo Minis", "$2.99", "", "", ""), ("Chips Ahoy Big Chewy Caramel", "$2.99", "", "", "Big Bag"),
    ("Chips Ahoy Mini", "$2.99", "", "", ""), ("Chips Ahoy Chewy", "$2.99", "", "", ""),
    ("Grandma's Cookies", "$1.69", "$2.99", "Big", ""), ("Keebler Sugar Wafers", "$2.99", "", "", "King Size, multiple flavors"),
    ("Keebler Fudge Stripes", "$2.99", "", "", "King Size"), ("Reese's Dipped Animal Crackers", "$2.99", "", "", ""),
    ("Hershey's Cookies n Creme Dipped", "$2.99", "", "", "Pretzels"), ("Reese's Dipped Pretzels", "$2.99", "", "", ""),
    ("Nutter Butter", "$1.69", "", "", ""), ("White Cloud Mini Cookies", "$1.69", "", "", ""),
]

candy_data = [
    ("Candy (generic single)", "$0.59", "$1.00", "2 for", "Various brands"), ("Reese's Cups", "$0.99", "$1.79", "2 for", ""),
    ("Snickers", "$0.99", "", "", ""), ("M&Ms", "$0.99", "", "", ""), ("Skittles", "$0.99", "", "", ""),
    ("Starburst", "$0.99", "", "", ""), ("Now and Laters", "$0.25", "$1.00", "5 for", ""),
    ("Jolly Ranchers (bag)", "$0.99", "", "", ""), ("Blow Pops", "$0.25", "", "", ""), ("Tootsie Rolls", "$0.25", "", "", ""),
    ("Honey Bun", "$0.99", "", "", ""), ("Oatmeal Cream Pie", "$0.99", "", "", ""),
    ("Swiss Rolls", "$0.99", "", "", ""), ("Zebra Cakes", "$0.99", "", "", ""),
]

popcorn_data = [("Act II Butter", "$1.59", "", "", "Microwave"), ("Act II Butter Lovers", "$1.59", "", "", ""),
                ("Popcorn (clubs & cheese)", "$1.19", "", "", ""), ("Sprinkles Popcorn Large", "$3.99", "$2.49", "Small", "")]
pickle_data = [("Pickle in a Pouch", "$1.99", "", "", "Kosher Dill"), ("Suckerpunch Pickle Snack Pack", "$1.99", "", "", ""),
               ("Vlasic Pickle Balls", "$2.99", "", "", ""), ("Pickles (jar)", "$2.49", "", "", "")]

cur5 = 4
cur5 = add_section(ws5, cur5, "GOLDEN FLAKE (Local Brand)", COLS5, gf_data)
cur5 = add_section(ws5, cur5, "CHIPS & SAVORY SNACKS", COLS5, chips_data)
cur5 = add_section(ws5, cur5, "COOKIES & CRACKERS", COLS5, cookies_data)
cur5 = add_section(ws5, cur5, "CANDY", COLS5, candy_data)
cur5 = add_section(ws5, cur5, "POPCORN", COLS5, popcorn_data)
cur5 = add_section(ws5, cur5, "PICKLES & PICKLED SNACKS", COLS5, pickle_data)
ws5.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 6 – BEVERAGES (Non-Alcoholic)
# ════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("🥤 Beverages")
COLS6 = 5
write_title(ws6, "BEVERAGES (NON-ALCOHOLIC)", "Sodas | Water | Juice | Energy | Sports", COLS6)
set_col_widths(ws6, [32, 10, 10, 12, 22])
hdrs6 = ["Product", "Price", "Deal Price", "Deal Qty", "Notes"]
for j, h in enumerate(hdrs6, 1):
    ws6.cell(row=3, column=j, value=h)
style_header_row(ws6, 3, COLS6)
ws6.row_dimensions[3].height = 22

faygo_data = [("Faygo Can (any flavor)", "$0.59", "$1.35", "3 for", ""), ("Faygo 2-Liter", "$1.49", "", "", ""), ("Faygo 20oz Bottle", "$1.49", "", "", "")]
pepsi_data = [("Pepsi 20oz", "$1.99", "", "", ""), ("Pepsi 2-Liter", "$2.49", "", "", ""), ("Diet Pepsi 20oz", "$1.99", "", "", ""),
              ("Mountain Dew 20oz", "$1.99", "", "", ""), ("Mountain Dew Code Red 20oz", "$1.99", "", "", ""), ("Mug Root Beer 20oz", "$1.99", "", "", ""),
              ("Lipton Brisk Tea 20oz", "$1.49", "", "", ""), ("Gatorade 20oz", "$1.99", "", "", "Various flavors"), ("Gatorade 32oz", "$2.49", "", "", ""),
              ("Tropicana Orange Juice", "$3.99", "$2.99", "", ""), ("Tropicana (small)", "$1.49", "", "", "")]
coke_data = [("Coca-Cola 20oz", "$1.99", "", "", ""), ("Coca-Cola 2-Liter", "$2.49", "", "", ""), ("Diet Coke 20oz", "$1.99", "", "", ""),
             ("Sprite 20oz", "$1.99", "", "", ""), ("Fanta Orange 20oz", "$1.99", "", "", ""), ("Dr Pepper 20oz", "$1.99", "", "", ""),
             ("Minute Maid Juice", "$2.99", "", "", "Loose/singles"), ("Powerade 20oz", "$1.49", "", "", ""), ("Body Armor", "$2.49", "", "", "")]
water_data = [("Water Small Bottle", "$0.99", "", "", ""), ("Water Large Bottle", "$1.49", "", "", ""), ("Dasani 20oz", "$1.99", "", "", ""),
              ("Smart Water", "$2.49", "", "", "")]
energy_data = [("Red Bull 8.4oz", "$2.99", "", "", ""), ("Red Bull 12oz", "$3.49", "", "", ""), ("Monster Energy 16oz", "$2.99", "", "", ""),
               ("Bang Energy 16oz", "$2.99", "", "", ""), ("5-Hour Energy Shot", "$2.99", "", "", ""), ("METAx", "$2.59", "", "", "")]
juice_data = [("Nectar Juice", "$2.99", "", "", ""), ("Orange Juice (small)", "$3.99", "", "", ""), ("Apple Juice", "$2.49", "", "", ""),
              ("Minute Maid Lemonade", "$1.99", "", "", "20oz")]

cur6 = 4
cur6 = add_section(ws6, cur6, "FAYGO (Local Favorite)", COLS6, faygo_data)
cur6 = add_section(ws6, cur6, "PEPSI PRODUCTS", COLS6, pepsi_data)
cur6 = add_section(ws6, cur6, "COCA-COLA PRODUCTS", COLS6, coke_data)
cur6 = add_section(ws6, cur6, "WATER", COLS6, water_data)
cur6 = add_section(ws6, cur6, "ENERGY DRINKS", COLS6, energy_data)
cur6 = add_section(ws6, cur6, "JUICE", COLS6, juice_data)
ws6.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 7 – ICE CREAM
# ════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("🍦 Ice Cream")
COLS7 = 4
write_title(ws7, "ICE CREAM & FROZEN TREATS", "Singles | Bars | Cups", COLS7)
set_col_widths(ws7, [34, 10, 12, 22])
hdrs7 = ["Product", "Price", "Deal Price", "Notes"]
for j, h in enumerate(hdrs7, 1):
    ws7.cell(row=3, column=j, value=h)
style_header_row(ws7, 3, COLS7)
ws7.row_dimensions[3].height = 22

ice_data = [
    ("Popsicle Firecracker", "$1.49", "", ""), ("Popsicle (standard)", "$1.49", "", ""), ("Screwball (candy-filled)", "$1.99", "", ""),
    ("Big Stick", "$1.49", "", ""), ("Creamsicle", "$1.49", "", ""), ("Ice Cream Sandwich", "$1.99", "", ""),
    ("King Cone / Drumstick", "$2.49", "", ""), ("Choco Taco", "$2.49", "", ""), ("Ben & Jerry's Cup", "$4.99", "", "Individual"),
    ("Haagen-Dazs Cup", "$3.99", "", "Individual"), ("Klondike Bar", "$1.99", "", ""), ("Magnum Bar", "$2.99", "", ""),
    ("Fudgesicle", "$1.49", "", ""), ("Push Pop", "$1.49", "", ""), ("Snow Cone Cup", "$1.99", "", ""),
    ("Bomb Pop", "$1.99", "", ""),
]

cur7 = 4
cur7 = add_section(ws7, cur7, "ICE CREAM & FROZEN TREATS", COLS7, ice_data)
ws7.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 8 – MEDICINE & PERSONAL CARE
# ════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("💊 Medicine & Care")
COLS8 = 4
write_title(ws8, "MEDICINE & PERSONAL CARE", "OTC Medicine | Personal Care | Hygiene", COLS8)
set_col_widths(ws8, [34, 10, 12, 22])
hdrs8 = ["Product", "Price", "Notes", ""]
for j, h in enumerate(hdrs8, 1):
    ws8.cell(row=3, column=j, value=h)
style_header_row(ws8, 3, COLS8)
ws8.row_dimensions[3].height = 22

med_data = [
    ("Goody's Headache Powder", "$0.59", "Classic formula", ""), ("Goody's Cool Orange", "$0.59", "", ""),
    ("Goody's Extra Strength", "$0.79", "", ""), ("BC Powder Original", "$0.59", "Aspirin blend", ""),
    ("BC Powder Arthritis", "$0.79", "", ""), ("Tylenol Extra Strength (2-pk)", "$1.49", "", ""),
    ("Tylenol Regular Strength", "$1.49", "", ""), ("Advil (2-pk)", "$1.49", "Ibuprofen", ""),
    ("Aleve (2-pk)", "$1.49", "Naproxen", ""), ("Aspirin (small pack)", "$0.99", "", ""),
    ("Tums Antacid", "$1.49", "", ""), ("Rolaids", "$1.49", "", ""), ("Pepto-Bismol (chews)", "$1.99", "", ""),
    ("Pepcid AC", "$1.99", "", ""), ("DayQuil (2-pk)", "$2.49", "", ""), ("NyQuil (2-pk)", "$2.49", "", ""),
    ("Benadryl (2-pk)", "$1.99", "", ""), ("Imodium AD", "$1.99", "", ""), ("Eye Drops (Visine)", "$2.99", "", ""),
    ("Cough Drops (Halls)", "$1.49", "", ""),
]

care_data = [
    ("Isopropyl Alcohol 70%", "$2.49", "", ""), ("Hydrogen Peroxide", "$1.49", "", ""), ("Band-Aids (small pack)", "$1.99", "", ""),
    ("Condoms (3-pack)", "$3.99", "", ""), ("Always Pads", "$4.59", "", ""), ("Tampax", "$4.59", "", ""),
    ("Facial Tissue", "$4.59", "", ""), ("Toothbrush", "$1.99", "", ""), ("Toothpaste (travel)", "$1.99", "", ""),
    ("Deodorant (small)", "$3.49", "", ""), ("Lotion (small)", "$2.99", "", ""), ("Lip Balm", "$1.49", "", ""),
    ("Body Armor Tropical Punch", "$2.49", "", "Sports drink"),
]

cur8 = 4
cur8 = add_section(ws8, cur8, "PAIN RELIEF & MEDICINE", COLS8, med_data)
cur8 = add_section(ws8, cur8, "PERSONAL CARE & HYGIENE", COLS8, care_data)
ws8.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 9 – HOUSEHOLD / FOOD
# ════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("🏠 Household & Food")
COLS9 = 4
write_title(ws9, "HOUSEHOLD SUPPLIES & HOT FOOD", "Kitchen | Condiments | Deli | Dairy", COLS9)
set_col_widths(ws9, [34, 10, 12, 22])
hdrs9 = ["Product", "Price", "Alt Price", "Notes"]
for j, h in enumerate(hdrs9, 1):
    ws9.cell(row=3, column=j, value=h)
style_header_row(ws9, 3, COLS9)
ws9.row_dimensions[3].height = 22

kitchen_data = [
    ("Paper Plates (4-pack)", "$0.65", "", ""), ("Paper Plates (30-pack)", "$2.99", "", ""), ("Plastic Cups", "$3.49", "", ""),
    ("Plastic Spoons & Forks", "$1.49", "", ""), ("Plastic Wrap / Handi Wrap", "$3.99", "", ""), ("Plastic Straws", "$1.89", "", ""),
    ("Plastic Napkins", "$2.99", "", ""), ("Plastic Bowls (4-pack)", "$3.99", "", ""), ("Plastic Tova Napkins", "$2.99", "", ""),
    ("Aluminum Foil", "$2.99", "", ""), ("Tissue (4-pack)", "$1.09", "", ""), ("Tissue Paper Plate 4-pack", "$2.49", "", ""),
    ("Zip-loc Bags", "$1.99", "", "Sandwich size"),
]

condiment_data = [("BBQ Sauce", "$2.99", "", ""), ("Ketchup", "$2.99", "", ""), ("Mustard", "$1.69", "", ""),
                  ("Hot Sauce (small)", "$1.49", "", ""), ("Soy Sauce Packet", "$0.25", "", ""), ("Mayo Packet", "$0.25", "", ""),
                  ("Salt & Pepper Packets", "$0.10", "", ""), ("Seasonings (small)", "$1.59", "", "")]
dairy_data = [("Milk (small)", "$3.49", "$2.49", "Prices vary"), ("Milk (regular)", "$2.99", "", ""), ("Eggs", "$3.49", "", ""),
              ("Butter (small)", "$2.49", "", ""), ("Yogurt", "$3.99", "", ""), ("Cheese Slices", "$2.49", "", ""),
              ("Orange Juice (carton)", "$3.99", "", ""), ("Nectar Juice", "$2.99", "", "")]
food_data = [("Bologna (deli)", "$2.49", "", "Per serving"), ("Burger", "$3.49", "", "Hot food"), ("Chicken (piece)", "$3.49", "", "Hot food"),
             ("Corn Dog", "$3.49", "", ""), ("Deli Express Sandwich", "$1.29", "", ""), ("Lunchable", "$3.99", "", ""),
             ("Philly Cheesesteak", "$2.99", "", "Hot/frozen"), ("Quesadilla", "$3.59", "", "Hot food"), ("White Castle (frozen 2-pack)", "$2.49", "", ""),
             ("Hot Dog", "$1.99", "", ""), ("Ramen Noodles", "$0.49", "$1.00", "2 for"), ("Cup Noodles", "$0.99", "", "")]

cur9 = 4
cur9 = add_section(ws9, cur9, "KITCHEN & DISPOSABLES", COLS9, kitchen_data)
cur9 = add_section(ws9, cur9, "CONDIMENTS & SAUCES", COLS9, condiment_data)
cur9 = add_section(ws9, cur9, "DAIRY & REFRIGERATED", COLS9, dairy_data)
cur9 = add_section(ws9, cur9, "HOT FOOD & DELI", COLS9, food_data)
ws9.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 10 – QUICK LOOKUP (Index)
# ════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("📋 Quick Lookup", 0)
COLS10 = 3
write_title(ws10, "⛽  HOOD GAS STATION – MASTER PRICE GUIDE", "Quick Reference Index | Use sheet tabs to navigate categories", COLS10)
set_col_widths(ws10, [30, 16, 30])
ws10.row_dimensions[1].height = 32
ws10.row_dimensions[2].height = 18

index_items = [
    ("🍺 Beer", "Beer sheet", "16oz | 4pk | 6pk | 12pk | 24pk | Single 24oz"),
    ("🍷 Wine & Malt", "Wine sheet", "Wine by bottle + Malt beverages"),
    ("🥃 Liquor", "Liquor sheet", "Vodka | Whiskey | Cognac | Rum | Tequila | Gin"),
    ("🚬 Tobacco", "Tobacco sheet", "Cigarettes | Cigars | Wraps | Dip"),
    ("🍿 Snacks", "Snacks sheet", "Golden Flake | Chips | Cookies | Candy | Pickles"),
    ("🥤 Beverages", "Beverages sheet", "Faygo | Pepsi | Coke | Water | Energy | Juice"),
    ("🍦 Ice Cream", "Ice Cream sheet", "Bars | Cups | Cones | Frozen Treats"),
    ("💊 Medicine", "Medicine sheet", "Goody's | Tylenol | BC | Personal Care"),
    ("🏠 Household", "Household sheet", "Supplies | Condiments | Dairy | Hot Food"),
]

index_hdrs = ["Category", "Sheet Tab", "Includes"]
for j, h in enumerate(index_hdrs, 1):
    ws10.cell(row=3, column=j, value=h)
style_header_row(ws10, 3, COLS10)
ws10.row_dimensions[3].height = 22

special_offers = [
    "",
    "══════════════════════════════════",
    "🏷️  CURRENT DEALS & SPECIALS",
    "══════════════════════════════════",
    "Faygo Can:         $0.59 each | 3 for $1.35",
    "Candy:             $0.59 each | 2 for $1.00",
    "Golden Flake:      $0.65 small bags",
    "Swisher Sweets:    2 for $1.00 (some varieties)",
    "Ramen Noodles:     2 for $1.00",
    "Beer 8-Pack Bud Ice: $9.99",
    "Natty Light 15pk:  $12.99",
    "",
]

for i, item in enumerate(index_items, 4):
    for j, val in enumerate(item, 1):
        ws10.cell(row=i, column=j, value=val)
    style_data_row(ws10, i, COLS10, alt=(i % 2 == 0))
    ws10.row_dimensions[i].height = 18

r = 4 + len(index_items) + 1
for line in special_offers:
    ws10.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS10)
    c = ws10.cell(row=r, column=1, value=line)
    c.font = Font(name="Arial", size=10, bold=("DEAL" in line or "══" in line or "CURRENT" in line))
    c.fill = fill(GOLD if "$" in line and not line.startswith("═") else WHITE)
    c.alignment = left()
    ws10.row_dimensions[r].height = 16
    r += 1

# ─── Save the file ──────────────────────────────────────────────────
out = "GasStation_PriceList.xlsx"
wb.save(out)
print(f"✅ Excel file saved as: {out}")
print(f"   Total sheets: {len(wb.sheetnames)} | All products listed exhaustively.")